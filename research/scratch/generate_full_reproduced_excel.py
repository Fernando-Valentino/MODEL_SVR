import os
import openpyxl
import pandas as pd
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sklearn.preprocessing import RobustScaler, MinMaxScaler
from sklearn.svm import SVR
from sklearn.metrics import r2_score

# 80 Combinations of Grid Search from skripsi/system results
GRID_SEARCH_DATA = [
    (1, 100, 0.001, 0.01, 0.258671, 0.029156, 0.258671, "Terbaik"),
    (2, 150, 0.001, 0.01, 0.259302, 0.028725, 0.259302, ""),
    (3, 200, 0.001, 0.01, 0.259548, 0.027784, 0.259548, ""),
    (4, 50, 0.001, 0.01, 0.260321, 0.031461, 0.260321, ""),
    (5, 10, 0.001, 0.05, 0.261114, 0.026071, 0.261114, ""),
    (6, 100, 0.005, 0.01, 0.261235, 0.030845, 0.261235, ""),
    (7, 10, 0.005, 0.05, 0.261265, 0.027991, 0.261265, ""),
    (8, 150, 0.005, 0.01, 0.261886, 0.031365, 0.261886, ""),
    (9, 50, 0.005, 0.01, 0.262003, 0.032659, 0.262003, ""),
    (10, 200, 0.005, 0.01, 0.262331, 0.031678, 0.262331, ""),
    (11, 10, 0.01, "scale", 0.262460, 0.026665, 0.262460, ""),
    (12, 50, 0.01, 0.01, 0.262635, 0.033977, 0.262635, ""),
    (13, 10, 0.001, "scale", 0.262724, 0.025450, 0.262724, ""),
    (14, 10, 0.005, "scale", 0.262868, 0.027158, 0.262868, ""),
    (15, 10, 0.01, 0.05, 0.263116, 0.028956, 0.263116, ""),
    (16, 100, 0.01, 0.01, 0.263942, 0.036215, 0.263942, ""),
    (17, 150, 0.01, 0.01, 0.264698, 0.036256, 0.264698, ""),
    (18, 200, 0.01, 0.01, 0.265298, 0.036575, 0.265298, ""),
    (19, 50, 0.01, 0.05, 0.267832, 0.022414, 0.267832, ""),
    (20, 50, 0.005, 0.05, 0.269739, 0.023187, 0.269739, ""),
    (21, 10, 0.01, 0.01, 0.269889, 0.036666, 0.269889, ""),
    (22, 50, 0.01, "scale", 0.270813, 0.023469, 0.270813, ""),
    (23, 10, 0.001, 0.01, 0.270840, 0.036395, 0.270840, ""),
    (24, 10, 0.005, 0.01, 0.270916, 0.036404, 0.270916, ""),
    (25, 50, 0.005, "scale", 0.271632, 0.023540, 0.271632, ""),
    (26, 50, 0.001, 0.05, 0.274870, 0.023691, 0.274870, ""),
    (27, 10, 0.05, "scale", 0.276261, 0.021088, 0.276261, ""),
    (28, 10, 0.05, 0.05, 0.276751, 0.022645, 0.276751, ""),
    (29, 50, 0.001, "scale", 0.277336, 0.023280, 0.277336, ""),
    (30, 100, 0.01, 0.05, 0.277470, 0.024721, 0.277470, ""),
    (31, 100, 0.005, 0.05, 0.278885, 0.024885, 0.278885, ""),
    (32, 200, 0.01, 0.001, 0.281698, 0.041466, 0.281698, ""),
    (33, 200, 0.005, 0.001, 0.281950, 0.041392, 0.281950, ""),
    (34, 200, 0.001, 0.001, 0.281973, 0.041159, 0.281973, ""),
    (35, 150, 0.01, 0.001, 0.282449, 0.041686, 0.282449, ""),
    (36, 50, 0.05, 0.05, 0.282477, 0.017900, 0.282477, ""),
    (37, 10, 0.05, 0.01, 0.282822, 0.029240, 0.282822, ""),
    (38, 150, 0.001, 0.001, 0.283032, 0.041319, 0.283032, ""),
    (39, 150, 0.005, 0.001, 0.283451, 0.041477, 0.283451, ""),
    (40, 100, 0.01, "scale", 0.283555, 0.024800, 0.283555, ""),
    (41, 100, 0.005, "scale", 0.283952, 0.026710, 0.283952, ""),
    (42, 100, 0.01, 0.001, 0.283959, 0.042056, 0.283959, ""),
    (43, 100, 0.001, 0.001, 0.284680, 0.042003, 0.284680, ""),
    (44, 100, 0.005, 0.001, 0.284932, 0.042009, 0.284932, ""),
    (45, 100, 0.001, 0.05, 0.285728, 0.025074, 0.285728, ""),
    (46, 50, 0.05, 0.01, 0.285974, 0.037081, 0.285974, ""),
    (47, 50, 0.01, 0.001, 0.286204, 0.042865, 0.286204, ""),
    (48, 150, 0.01, 0.05, 0.286311, 0.025759, 0.286311, ""),
    (49, 50, 0.05, "scale", 0.286402, 0.018653, 0.286402, ""),
    (50, 100, 0.05, 0.001, 0.286785, 0.029942, 0.286785, ""),
    (51, 150, 0.005, 0.05, 0.287171, 0.026786, 0.287171, ""),
    (52, 50, 0.005, 0.001, 0.287375, 0.043479, 0.287375, ""),
    (53, 50, 0.001, 0.001, 0.287547, 0.043406, 0.287547, ""),
    (54, 150, 0.05, 0.001, 0.288269, 0.030960, 0.288269, ""),
    (55, 50, 0.05, 0.001, 0.288769, 0.031251, 0.288769, ""),
    (56, 100, 0.001, "scale", 0.289120, 0.025489, 0.289120, ""),
    (57, 200, 0.05, 0.001, 0.291749, 0.035888, 0.291749, ""),
    (58, 150, 0.001, 0.05, 0.293506, 0.026248, 0.293506, ""),
    (59, 200, 0.05, 0.01, 0.293853, 0.045709, 0.293853, ""),
    (60, 100, 0.05, 0.05, 0.294263, 0.022069, 0.294263, ""),
    (61, 100, 0.05, 0.01, 0.294305, 0.046752, 0.294305, ""),
    (62, 150, 0.01, "scale", 0.294592, 0.029542, 0.294592, ""),
    (63, 200, 0.01, 0.05, 0.295235, 0.026649, 0.295235, ""),
    (64, 150, 0.05, 0.01, 0.295291, 0.047959, 0.295291, ""),
    (65, 10, 0.01, 0.001, 0.295303, 0.046398, 0.295303, ""),
    (66, 10, 0.05, 0.001, 0.295337, 0.037296, 0.295337, ""),
    (67, 200, 0.005, 0.05, 0.295489, 0.028253, 0.295489, ""),
    (68, 150, 0.005, "scale", 0.295925, 0.028637, 0.295925, ""),
    (69, 10, 0.005, 0.001, 0.296072, 0.047128, 0.296072, ""),
    (70, 10, 0.001, 0.001, 0.296566, 0.047369, 0.296566, ""),
    (71, 100, 0.05, "scale", 0.297284, 0.023074, 0.297284, ""),
    (72, 150, 0.001, "scale", 0.299818, 0.028433, 0.299818, ""),
    (73, 150, 0.05, 0.05, 0.300195, 0.024459, 0.300195, ""),
    (74, 200, 0.001, 0.05, 0.300838, 0.027826, 0.300838, ""),
    (75, 150, 0.05, "scale", 0.300985, 0.022723, 0.300985, ""),
    (76, 200, 0.05, "scale", 0.301014, 0.022665, 0.301014, ""),
    (77, 200, 0.05, 0.05, 0.302839, 0.023947, 0.302839, ""),
    (78, 200, 0.01, "scale", 0.304089, 0.037278, 0.304089, ""),
    (79, 200, 0.005, "scale", 0.305436, 0.036146, 0.305436, ""),
    (80, 200, 0.001, "scale", 0.310276, 0.033600, 0.310276, "")
]

# 17 Convergence logs of Grey Wolf Optimizer
GWO_CONVERGENCE_DATA = [
    (1, 0.067408, 250.034500, 0.005366, 0.004455, 0.070778, 0.072800, 0.067408, "Membaik", "—"),
    (2, 0.067408, 250.034500, 0.005366, 0.004455, 0.070778, 0.072800, 0.067408, "Stagnasi", "0.0000%"),
    (3, 0.067408, 250.034500, 0.005366, 0.004455, 0.070778, 0.072800, 0.067408, "Stagnasi", "0.0000%"),
    (4, 0.067408, 250.034500, 0.005366, 0.004455, 0.070778, 0.072800, 0.067408, "Stagnasi", "0.0000%"),
    (5, 0.067348, 250.034500, 0.005366, 0.004455, 0.070715, 0.072735, 0.067348, "Membaik", "0.0892%"),
    (6, 0.067348, 250.034500, 0.005366, 0.004455, 0.070715, 0.072735, 0.067348, "Stagnasi", "0.0000%"),
    (7, 0.067348, 250.034500, 0.005366, 0.004455, 0.070715, 0.072735, 0.067348, "Stagnasi", "0.0000%"),
    (8, 0.067348, 250.034500, 0.005366, 0.004455, 0.070715, 0.072735, 0.067348, "Stagnasi", "0.0000%"),
    (9, 0.067289, 250.034500, 0.005366, 0.004455, 0.070654, 0.072672, 0.067289, "Membaik", "0.0867%"),
    (10, 0.067289, 250.034500, 0.005366, 0.004455, 0.070654, 0.072672, 0.067289, "Stagnasi", "0.0000%"),
    (11, 0.067289, 250.034500, 0.005366, 0.004455, 0.070654, 0.072672, 0.067289, "Stagnasi", "0.0000%"),
    (12, 0.067289, 250.034500, 0.005366, 0.004455, 0.070654, 0.072672, 0.067289, "Stagnasi", "0.0000%"),
    (13, 0.067289, 250.034500, 0.005366, 0.004455, 0.070654, 0.072672, 0.067289, "Stagnasi", "0.0000%"),
    (14, 0.067289, 250.034500, 0.005366, 0.004455, 0.070654, 0.072672, 0.067289, "Stagnasi", "0.0000%"),
    (15, 0.067289, 250.034500, 0.005366, 0.004455, 0.070654, 0.072672, 0.067289, "Stagnasi", "0.0000%"),
    (16, 0.067289, 250.034500, 0.005366, 0.004455, 0.070654, 0.072672, 0.067289, "Stagnasi", "0.0000%"),
    (17, 0.067289, 250.034500, 0.005366, 0.004455, 0.070654, 0.072672, 0.067289, "Stagnasi", "0.0000%")
]

def export_to_excel_file(df_test, y_test_asli, y_pred_default, y_pred_gs, y_pred_gwo):
    print("Generating Reproduced Excel File...")
    wb = openpyxl.Workbook()
    
    # Styles Setup
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    accent_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    best_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # soft green
    alt_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True, color="000000")
    regular_font = Font(name="Calibri", size=11, color="000000")
    green_bold_font = Font(name="Calibri", size=11, bold=True, color="375623")
    
    # ════════════════════════════════════════════════════════════════════
    # SHEET 1 — PERBANDINGAN METRIK
    # ════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "1_Perbandingan_Metrik"
    ws1.sheet_view.showGridLines = True
    
    for col, w in [("A", 4), ("B", 24), ("C", 16), ("D", 20), ("E", 16), ("F", 4), ("G", 22), ("H", 16)]:
        ws1.column_dimensions[col].width = w
        
    ws1.merge_cells("B2:E2")
    title_cell = ws1["B2"]
    title_cell.value = "Perbandingan Tiga Model SVR Default vs SVR + Grid Search vs SVR + GWO"
    title_cell.font = Font(name="Calibri", size=12, bold=True, color="1B3A5C")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    
    headers = ["Metrik", "SVR Default", "SVR + Grid Search", "SVR + GWO"]
    for i, h in enumerate(headers, start=2):
        cell = ws1.cell(row=4, column=i, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws1.row_dimensions[4].height = 28
    
    rows_data = [
        ("─── Parameter ───", True, "", "", "", ""),
        ("C", False, 1.0, 100, 250.034536, "General"),
        ("Epsilon ε", False, 0.1, 0.001, 0.00536603, "General"),
        ("Gamma γ", False, "scale", 0.01, 0.004455, "General"),
        ("─── Performa Train ───", True, "", "", "", ""),
        ("MAE Train", False, 297195, 46364, 59889, "Rp #,##0"),
        ("RMSE Train", False, 357172, 119067, 132308, "Rp #,##0"),
        ("MAPE Train", False, 0.226831, 0.072972, 0.093622, "0.0000%"),
        ("R² Train", False, 0.642689, 0.960292, 0.950970, "0.000000"),
        ("─── Performa Test ───", True, "", "", "", ""),
        ("MAE Test", False, 369655, 135957, 130623, "Rp #,##0"),
        ("RMSE Test", False, 451420, 203896, 194009, "Rp #,##0"),
        ("MAPE Test", False, 0.251129, 0.130788, 0.129644, "0.0000%"),
        ("Gap Overfit", False, 0.024298, 0.057816, 0.036022, "0.0000%"),
        ("R² Test", False, 0.520081, 0.902091, 0.911356, "0.000000"),
        ("Akurasi Test", False, 0.7489, 0.8692, 0.8704, "0.00%"),
        ("Waktu Training", False, "0.2s", "1778.0s", "4293.8s", "General")
    ]
    
    current_row = 5
    for label, is_hdr, v_def, v_gs, v_gwo, num_fmt in rows_data:
        c_lbl = ws1.cell(row=current_row, column=2, value=label)
        c_lbl.border = thin_border
        
        c_def = ws1.cell(row=current_row, column=3, value=v_def)
        c_def.border = thin_border
        
        c_gs = ws1.cell(row=current_row, column=4, value=v_gs)
        c_gs.border = thin_border
        
        c_gwo = ws1.cell(row=current_row, column=5, value=v_gwo)
        c_gwo.border = thin_border
        
        if is_hdr:
            c_lbl.font = bold_font
            c_lbl.fill = accent_fill
            ws1.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=5)
            for col in range(2, 6):
                ws1.cell(row=current_row, column=col).border = thin_border
                ws1.cell(row=current_row, column=col).fill = accent_fill
        else:
            c_lbl.font = bold_font if "Train" in label or "Test" in label or "Overfit" in label else regular_font
            c_def.font = regular_font
            c_gs.font = regular_font
            c_gwo.font = green_bold_font if label == "Akurasi Test" or label == "MAPE Test" else regular_font
            
            c_def.alignment = Alignment(horizontal="right" if type(v_def) in [int, float] else "center")
            c_gs.alignment = Alignment(horizontal="right" if type(v_gs) in [int, float] else "center")
            c_gwo.alignment = Alignment(horizontal="right" if type(v_gwo) in [int, float] else "center")
            
            if num_fmt != "General" and num_fmt != "":
                c_def.number_format = num_fmt
                c_gs.number_format = num_fmt
                c_gwo.number_format = num_fmt
                
            if label in ["MAE Test", "RMSE Test", "MAPE Test", "R² Test", "Akurasi Test"]:
                c_gwo.fill = best_fill
                
        ws1.row_dimensions[current_row].height = 20
        current_row += 1
        
    # Best Model Summary Table on the side
    ws1.merge_cells("G4:H4")
    best_summary_hdr = ws1["G4"]
    best_summary_hdr.value = "Ringkasan Model Terbaik"
    best_summary_hdr.font = white_font
    best_summary_hdr.fill = header_fill
    best_summary_hdr.alignment = Alignment(horizontal="center", vertical="center")
    best_summary_hdr.border = thin_border
    ws1.cell(row=4, column=8).border = thin_border
    
    ws1.merge_cells("G5:H5")
    model_name_cell = ws1["G5"]
    model_name_cell.value = "Model Terbaik: SVR + GWO"
    model_name_cell.font = Font(name="Calibri", size=11, bold=True, color="375623")
    model_name_cell.fill = best_fill
    model_name_cell.alignment = Alignment(horizontal="center", vertical="center")
    model_name_cell.border = thin_border
    ws1.cell(row=5, column=8).border = thin_border
    ws1.cell(row=5, column=8).fill = best_fill
    
    best_rows = [
        ("MAE Test", 130623, "Rp #,##0"),
        ("RMSE Test", 194009, "Rp #,##0"),
        ("MAPE Test", 0.129644, "0.0000%"),
        ("R² Test", 0.911356, "0.000000"),
        ("Akurasi", 0.8704, "0.00%"),
        ("Waktu", "4293.8s", "General")
    ]
    
    side_row = 6
    for lbl, val, num_fmt in best_rows:
        c_lbl = ws1.cell(row=side_row, column=7, value=lbl)
        c_lbl.font = bold_font
        c_lbl.border = thin_border
        c_lbl.fill = accent_fill
        c_lbl.alignment = Alignment(horizontal="left", vertical="center")
        
        c_val = ws1.cell(row=side_row, column=8, value=val)
        c_val.font = green_bold_font if lbl in ["Akurasi", "MAPE Test"] else regular_font
        c_val.border = thin_border
        c_val.alignment = Alignment(horizontal="right" if type(val) in [int, float] else "center")
        
        if num_fmt != "General":
            c_val.number_format = num_fmt
            
        side_row += 1

    # ════════════════════════════════════════════════════════════════════
    # SHEET 2 — HASIL PREDIKSI TEST SET
    # ════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("2_Hasil_Prediksi_Test")
    ws2.sheet_view.showGridLines = True
    
    for col, w in [
        ("A", 4), ("B", 8), ("C", 14), ("D", 8), ("E", 16),
        ("F", 18), ("G", 18), ("H", 18), ("I", 18), ("J", 18),
        ("K", 18), ("L", 18), ("M", 18), ("N", 18)
    ]:
        ws2.column_dimensions[col].width = w
        
    # Title
    ws2.merge_cells("B2:N2")
    ws2["B2"] = "Hasil Prediksi - 765 Data Test  |  Aktual vs SVR Default vs Grid Search vs GWO"
    ws2["B2"].font = Font(name="Calibri", size=12, bold=True, color="1B3A5C")
    ws2["B2"].alignment = Alignment(horizontal="left", vertical="center")
    
    # Headers
    # Identifikasi / Target / SVR Default / SVR + Grid Search / SVR + GWO
    ws2.merge_cells("B4:D4")
    ws2["B4"] = "Identifikasi"
    ws2["B4"].font = white_font
    ws2["B4"].fill = header_fill
    ws2["B4"].alignment = Alignment(horizontal="center", vertical="center")
    ws2["B4"].border = thin_border
    for c in range(2, 5): ws2.cell(row=4, column=c).border = thin_border
    
    ws2.merge_cells("E4:E4")
    ws2["E4"] = "Target"
    ws2["E4"].font = white_font
    ws2["E4"].fill = header_fill
    ws2["E4"].alignment = Alignment(horizontal="center", vertical="center")
    ws2["E4"].border = thin_border
    
    ws2.merge_cells("F4:H4")
    ws2["F4"] = "SVR Default"
    ws2["F4"].font = white_font
    ws2["F4"].fill = header_fill
    ws2["F4"].alignment = Alignment(horizontal="center", vertical="center")
    ws2["F4"].border = thin_border
    for c in range(6, 9): ws2.cell(row=4, column=c).border = thin_border
    
    ws2.merge_cells("I4:K4")
    ws2["I4"] = "SVR + Grid Search"
    ws2["I4"].font = white_font
    ws2["I4"].fill = header_fill
    ws2["I4"].alignment = Alignment(horizontal="center", vertical="center")
    ws2["I4"].border = thin_border
    for c in range(9, 12): ws2.cell(row=4, column=c).border = thin_border
    
    ws2.merge_cells("L4:N4")
    ws2["L4"] = "SVR + GWO"
    ws2["L4"].font = white_font
    ws2["L4"].fill = header_fill
    ws2["L4"].alignment = Alignment(horizontal="center", vertical="center")
    ws2["L4"].border = thin_border
    for c in range(12, 15): ws2.cell(row=4, column=c).border = thin_border
    ws2.row_dimensions[4].height = 24
    
    subheaders = [
        "No", "Tanggal", "Rayon", "Aktual (Rp)",
        "SVR Default (Rp)", "Error Default", "MAPE Default%",
        "Grid Search (Rp)", "Error GS", "MAPE GS%",
        "GWO (Rp)", "Error GWO", "MAPE GWO%"
    ]
    for idx, sh in enumerate(subheaders, start=2):
        cell = ws2.cell(row=5, column=idx, value=sh)
        cell.font = white_font
        cell.fill = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws2.row_dimensions[5].height = 24
    
    # Populate rows
    n_records = len(y_test_asli)
    for r_idx in range(n_records):
        row_num = r_idx + 6
        bg_row = alt_fill if r_idx % 2 == 0 else PatternFill(fill_type=None)
        
        # Extract row info
        tgl = str(df_test.iloc[r_idx]['Tanggal'].date()) if r_idx < len(df_test) else "—"
        rayon = int(df_test.iloc[r_idx]['Rayon_asli']) if r_idx < len(df_test) else int(df_test.iloc[r_idx]['Rayon']) if 'Rayon' in df_test.columns else "—"
        
        act = float(y_test_asli[r_idx])
        pred_def = float(y_pred_default[r_idx])
        pred_gs = float(y_pred_gs[r_idx])
        pred_gwo = float(y_pred_gwo[r_idx])
        
        err_def = pred_def - act
        err_gs = pred_gs - act
        err_gwo = pred_gwo - act
        
        mape_def = abs(err_def) / max(act, 1.0)
        mape_gs = abs(err_gs) / max(act, 1.0)
        mape_gwo = abs(err_gwo) / max(act, 1.0)
        
        # Insert cells
        vals = [
            r_idx + 1, tgl, rayon, act,
            pred_def, err_def, mape_def,
            pred_gs, err_gs, mape_gs,
            pred_gwo, err_gwo, mape_gwo
        ]
        
        for col_idx, val in enumerate(vals, start=2):
            cell = ws2.cell(row=row_num, column=col_idx, value=val)
            cell.border = thin_border
            if bg_row.fill_type:
                cell.fill = bg_row
                
            cell.font = regular_font
            
            # Formatting
            if col_idx in [2, 3, 4]: # No, Tanggal, Rayon
                cell.alignment = Alignment(horizontal="center")
            else: # Currency or Percentage
                cell.alignment = Alignment(horizontal="right")
                
            if col_idx in [5, 6, 7, 9, 10, 12, 13]: # Currency
                cell.number_format = "#,##0"
            elif col_idx in [8, 11, 14]: # Percentage APE
                cell.number_format = "0.0000%"
                
            # Soft green background for cells where GWO has lower error than GS and Default
            if col_idx in [12, 13, 14] and abs(err_gwo) < abs(err_gs) and abs(err_gwo) < abs(err_def):
                cell.fill = best_fill
                
        ws2.row_dimensions[row_num].height = 20

    # ════════════════════════════════════════════════════════════════════
    # SHEET 3 — GRID SEARCH COMBINATIONS
    # ════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("3_Kombinasi_Grid_Search")
    ws3.sheet_view.showGridLines = True
    
    for col, w in [("A", 4), ("B", 8), ("C", 12), ("D", 14), ("E", 14), ("F", 18), ("G", 18), ("H", 18), ("I", 18)]:
        ws3.column_dimensions[col].width = w
        
    # Title
    ws3.merge_cells("B2:I2")
    ws3["B2"] = "Grid Search - 80 Kombinasi × 5 Fold  |  Best: C=100  ε=0.001  γ=0.01"
    ws3["B2"].font = Font(name="Calibri", size=12, bold=True, color="1B3A5C")
    ws3["B2"].alignment = Alignment(horizontal="left", vertical="center")
    
    # Headers
    headers_gs = ["Rank", "C", "Epsilon ε", "Gamma γ", "Avg RMSE CV", "Std RMSE", "RMSE", "Status"]
    for idx, h in enumerate(headers_gs, start=2):
        cell = ws3.cell(row=4, column=idx, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws3.row_dimensions[4].height = 28
    
    # Populate rows
    for r_idx, (rank, C_val, eps_val, gamma_val, avg_rmse, std_rmse, rmse, status) in enumerate(GRID_SEARCH_DATA):
        row_num = r_idx + 5
        bg_row = alt_fill if r_idx % 2 == 0 else PatternFill(fill_type=None)
        
        vals = [rank, C_val, eps_val, gamma_val, avg_rmse, std_rmse, rmse, status]
        for col_idx, val in enumerate(vals, start=2):
            cell = ws3.cell(row=row_num, column=col_idx, value=val)
            cell.border = thin_border
            if bg_row.fill_type:
                cell.fill = bg_row
                
            cell.font = regular_font
            
            # Alignments
            if col_idx in [2, 3, 4, 5, 9]: # Rank, C, eps, gamma, status
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="right")
                
            # Number formats
            if col_idx in [6, 7, 8]:
                cell.number_format = "0.000000"
                
            # Best Row coloring
            if rank == 1:
                cell.fill = best_fill
                cell.font = Font(name="Calibri", size=11, bold=True, color="375623")
                
        ws3.row_dimensions[row_num].height = 20

    # ════════════════════════════════════════════════════════════════════
    # SHEET 4 — GWO CONVERGENCE LOGS
    # ════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("4_Log_Konvergensi_GWO")
    ws4.sheet_view.showGridLines = True
    
    for col, w in [
        ("A", 4), ("B", 10), ("C", 14), ("D", 14), ("E", 14),
        ("F", 14), ("G", 14), ("H", 14), ("I", 18), ("J", 14), ("K", 14)
    ]:
        ws4.column_dimensions[col].width = w
        
    # Title
    ws4.merge_cells("B2:K2")
    ws4["B2"] = "Grey Wolf Optimizer - 17 Iterasi  |  Best: C=250.0345  ε=0.005366  γ=0.004455"
    ws4["B2"].font = Font(name="Calibri", size=12, bold=True, color="1B3A5C")
    ws4["B2"].alignment = Alignment(horizontal="left", vertical="center")
    
    # Headers
    headers_gwo = ["Iterasi", "α RMSE", "α C", "α ε", "α γ", "β RMSE", "δ RMSE", "Best RMSE", "Status", "Δ (%)"]
    for idx, h in enumerate(headers_gwo, start=2):
        cell = ws4.cell(row=4, column=idx, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws4.row_dimensions[4].height = 28
    
    # Populate rows
    for r_idx, (iter_num, alpha_rmse, alpha_c, alpha_eps, alpha_gam, beta_rmse, delta_rmse, best_rmse, status, delta_pct) in enumerate(GWO_CONVERGENCE_DATA):
        row_num = r_idx + 5
        bg_row = alt_fill if r_idx % 2 == 0 else PatternFill(fill_type=None)
        
        vals = [iter_num, alpha_rmse, alpha_c, alpha_eps, alpha_gam, beta_rmse, delta_rmse, best_rmse, status, delta_pct]
        for col_idx, val in enumerate(vals, start=2):
            cell = ws4.cell(row=row_num, column=col_idx, value=val)
            cell.border = thin_border
            if bg_row.fill_type:
                cell.fill = bg_row
                
            cell.font = regular_font
            
            # Alignments
            if col_idx in [2, 10, 11]: # Iterasi, status, delta_pct
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="right")
                
            # Number formats
            if col_idx in [3, 7, 8, 9]: # RMSE values
                cell.number_format = "0.000000"
            elif col_idx in [4]: # C value
                cell.number_format = "#,##0.0000"
            elif col_idx in [5, 6]: # Epsilon and Gamma values
                cell.number_format = "0.000000"
                
            # Status colors
            if status == "Membaik":
                cell.fill = best_fill
                cell.font = Font(name="Calibri", size=11, bold=True, color="375623")
                
        ws4.row_dimensions[row_num].height = 20
        
    os.makedirs("research/reproduced_excel", exist_ok=True)
    wb.save("research/reproduced_excel/reproduced_svr_metrics.xlsx")
    print("Success! Excel file generated at research/reproduced_excel/reproduced_svr_metrics.xlsx")
