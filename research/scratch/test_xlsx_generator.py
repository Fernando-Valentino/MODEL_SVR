import os
import openpyxl
import pandas as pd
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def test_generate_excel():
    print("Testing Excel Generator...")
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "1_Perbandingan_Metrik"
    ws1.sheet_view.showGridLines = True
    
    # Border styles
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    accent_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    best_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # soft green
    
    white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True, color="000000")
    regular_font = Font(name="Calibri", size=11, color="000000")
    green_bold_font = Font(name="Calibri", size=11, bold=True, color="375623")
    
    # ════════════════════════════════════════════════════════════════════
    # SHEET 1 — PERBANDINGAN METRIK
    # ════════════════════════════════════════════════════════════════════
    # Setup columns width
    for col, w in [("A", 4), ("B", 24), ("C", 16), ("D", 20), ("E", 16), ("F", 4), ("G", 22), ("H", 16)]:
        ws1.column_dimensions[col].width = w
        
    # Title Block
    ws1.merge_cells("B2:E2")
    title_cell = ws1["B2"]
    title_cell.value = "Perbandingan Tiga Model SVR Default vs SVR + Grid Search vs SVR + GWO"
    title_cell.font = Font(name="Calibri", size=12, bold=True, color="1B3A5C")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Table Headers
    headers = ["Metrik", "SVR Default", "SVR + Grid Search", "SVR + GWO"]
    for i, h in enumerate(headers, start=2):
        cell = ws1.cell(row=4, column=i)
        cell.value = h
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws1.row_dimensions[4].height = 28
    
    # Rows definitions: (Label, is_header_row, v_default, v_gs, v_gwo, number_format)
    rows_data = [
        ("─── Parameter ───", True, "", "", "", ""),
        ("C", False, 1.0, 100, 250.034536, "General"),
        ("Epsilon ε", False, 0.1, 0.001, 0.00536603, "General"),
        ("Gamma γ", False, "scale", 0.01, 0.004455, "General"),
        ("─── Performa Train ───", True, "", "", "", ""),
        ("MAE Train", False, 297195, 46364, 59889, "Rp #,##0"),
        ("RMSE Train", False, 357172, 119067, 132308, "Rp #,##0"),
        ("MAPE Train", False, 0.226831, 0.072972, 0.093622, "0.0000%"),
        ("R² Train", False, 0.642689, 0.960292, 0.950970, "0.000000"),
        ("─── Performa Test ───", True, "", "", "", ""),
        ("MAE Test", False, 369655, 135957, 130623, "Rp #,##0"),
        ("RMSE Test", False, 451420, 203896, 194009, "Rp #,##0"),
        ("MAPE Test", False, 0.251129, 0.130788, 0.129644, "0.0000%"),
        ("Gap Overfit", False, 0.024298, 0.057816, 0.036022, "0.0000%"),
        ("R² Test", False, 0.520081, 0.902091, 0.911356, "0.000000"),
        ("Akurasi Test", False, 0.7489, 0.8692, 0.8704, "0.00%"),
        ("Waktu Training", False, "0.2s", "1778.0s", "4293.8s", "General")
    ]
    
    current_row = 5
    for label, is_hdr, v_def, v_gs, v_gwo, num_fmt in rows_data:
        # Write Label
        c_lbl = ws1.cell(row=current_row, column=2, value=label)
        c_lbl.border = thin_border
        
        c_def = ws1.cell(row=current_row, column=3, value=v_def)
        c_def.border = thin_border
        
        c_gs = ws1.cell(row=current_row, column=4, value=v_gs)
        c_gs.border = thin_border
        
        c_gwo = ws1.cell(row=current_row, column=5, value=v_gwo)
        c_gwo.border = thin_border
        
        if is_hdr:
            c_lbl.font = bold_font
            c_lbl.fill = accent_fill
            ws1.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=5)
            # Apply border to merged cells
            for col in range(2, 6):
                ws1.cell(row=current_row, column=col).border = thin_border
                ws1.cell(row=current_row, column=col).fill = accent_fill
        else:
            c_lbl.font = bold_font if "Train" in label or "Test" in label or "Overfit" in label else regular_font
            c_def.font = regular_font
            c_gs.font = regular_font
            c_gwo.font = green_bold_font if label == "Akurasi Test" or label == "MAPE Test" else regular_font
            
            c_def.alignment = Alignment(horizontal="right" if type(v_def) in [int, float] else "center")
            c_gs.alignment = Alignment(horizontal="right" if type(v_gs) in [int, float] else "center")
            c_gwo.alignment = Alignment(horizontal="right" if type(v_gwo) in [int, float] else "center")
            
            if num_fmt != "General" and num_fmt != "":
                c_def.number_format = num_fmt
                c_gs.number_format = num_fmt
                c_gwo.number_format = num_fmt
                
            # Apply soft green background to the best model (GWO) column for metrics
            if label in ["MAE Test", "RMSE Test", "MAPE Test", "R² Test", "Akurasi Test"]:
                c_gwo.fill = best_fill
                
        ws1.row_dimensions[current_row].height = 20
        current_row += 1
        
    # Best Model Summary Table on the side (starting column G)
    ws1.merge_cells("G4:H4")
    best_summary_hdr = ws1["G4"]
    best_summary_hdr.value = "Ringkasan Model Terbaik"
    best_summary_hdr.font = white_font
    best_summary_hdr.fill = header_fill
    best_summary_hdr.alignment = Alignment(horizontal="center", vertical="center")
    best_summary_hdr.border = thin_border
    ws1.cell(row=4, column=8).border = thin_border
    
    ws1.merge_cells("G5:H5")
    model_name_cell = ws1["G5"]
    model_name_cell.value = "Model Terbaik: SVR + GWO"
    model_name_cell.font = Font(name="Calibri", size=11, bold=True, color="375623")
    model_name_cell.fill = best_fill
    model_name_cell.alignment = Alignment(horizontal="center", vertical="center")
    model_name_cell.border = thin_border
    ws1.cell(row=5, column=8).border = thin_border
    ws1.cell(row=5, column=8).fill = best_fill
    
    best_rows = [
        ("MAE Test", 130623, "Rp #,##0"),
        ("RMSE Test", 194009, "Rp #,##0"),
        ("MAPE Test", 0.129644, "0.0000%"),
        ("R² Test", 0.911356, "0.000000"),
        ("Akurasi", 0.8704, "0.00%"),
        ("Waktu", "4293.8s", "General")
    ]
    
    side_row = 6
    for lbl, val, num_fmt in best_rows:
        c_lbl = ws1.cell(row=side_row, column=7, value=lbl)
        c_lbl.font = bold_font
        c_lbl.border = thin_border
        c_lbl.fill = accent_fill
        c_lbl.alignment = Alignment(horizontal="left", vertical="center")
        
        c_val = ws1.cell(row=side_row, column=8, value=val)
        c_val.font = green_bold_font if lbl in ["Akurasi", "MAPE Test"] else regular_font
        c_val.border = thin_border
        c_val.alignment = Alignment(horizontal="right" if type(val) in [int, float] else "center")
        
        if num_fmt != "General":
            c_val.number_format = num_fmt
            
        side_row += 1
        
    os.makedirs("research/reproduced_excel", exist_ok=True)
    wb.save("research/reproduced_excel/reproduced_svr_metrics.xlsx")
    print("Success! File saved.")

if __name__ == "__main__":
    test_generate_excel()
